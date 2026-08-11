# =========================
# IMPORTS
# =========================
import random
import numpy as np
import re
import time

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pymongo import MongoClient
from fastapi.responses import FileResponse

from aicte_pos import AICTE_POS
from db import get_cos_by_subject
from faculty_memory import store_feedback
from question_mapper import map_questions_to_co
from semantic_mapper import map_course_outcomes_to_pos, map_texts_to_candidates

# =========================
# FASTAPI APP
# =========================
from fastapi.middleware.cors import CORSMiddleware

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://ai-nlp-co-po-mapper-frontend.onrender.com",
        "http://localhost:3000",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================
# MONGODB CONNECTION
# =========================
import os
import bcrypt
from dotenv import load_dotenv
from pymongo import MongoClient

# Load environment variables
load_dotenv()

MONGODB_URI = os.getenv("MONGODB_URI", "mongodb://localhost:27017/?directConnection=true")
client = MongoClient(MONGODB_URI)

db = client["ai_copo_db"]

users_collection = db["signup"]
mapping_collection = db["mappings"]
YEAR_LABEL_TO_DB = {
    "FE": 1,
    "SE": 2,
    "TE": 3,
    "BE": 4,
    "1": 1,
    "2": 2,
    "3": 3,
    "4": 4,
}
ANALYTICS_SUBJECT_ALIASES = {
    "DBMS": ["dbms", "database management system", "database management systems"],
    "Operating System": ["operating system", "operating systems", "os"],
    "Computer Networks": ["computer networks", "computer network", "cn"],
    "Distributed Computing": ["distributed computing", "distributed systems"],
}
RUNTIME_CACHE_TTL_SECONDS = 180
_wk_catalog_cache = {"value": None, "expires_at": 0.0}
_level_threshold_cache = {"value": None, "expires_at": 0.0}


# =========================
# AICTE PROGRAM OUTCOMES
# =========================
# Imported from aicte_pos.py so the backend uses one source of truth.


def load_wk_catalog():
    now = time.monotonic()
    cached_value = _wk_catalog_cache.get("value")

    if cached_value is not None and now < _wk_catalog_cache.get("expires_at", 0.0):
        return [dict(item) for item in cached_value]

    wk_docs = list(
        db["WK"].find(
            {},
            {
                "_id": 0,
                "code": 1,
                "description": 1,
                "text": 1,
                "name": 1,
            },
        )
    )
    catalog = []

    for doc in wk_docs:
        code = (doc.get("code") or "").strip()
        text = (
            doc.get("description")
            or doc.get("text")
            or doc.get("name")
            or ""
        ).strip()

        if code and text:
            catalog.append({"code": code, "text": text})

    _wk_catalog_cache["value"] = [dict(item) for item in catalog]
    _wk_catalog_cache["expires_at"] = now + RUNTIME_CACHE_TTL_SECONDS
    return [dict(item) for item in _wk_catalog_cache["value"]]


def normalize_schema_value(schema):
    value = (schema or "").strip().lower()

    if value.startswith("c"):
        return "C"
    if value.startswith("n"):
        return "NEP"

    return "C"


def normalize_year_value(year):
    cleaned = str(year or "").strip().upper()
    return YEAR_LABEL_TO_DB.get(cleaned)


def get_subject_collection(schema):
    normalized_schema = normalize_schema_value(schema)
    return db["c_scheme"] if normalized_schema == "C" else db["nep_collection"]


def load_subject_catalog(schema, year=None, semester=None):
    collection = get_subject_collection(schema)
    query = {}
    normalized_year = normalize_year_value(year)

    if normalized_year is not None:
        query["year"] = normalized_year

    if semester not in (None, ""):
        try:
            query["semester"] = int(semester)
        except (TypeError, ValueError):
            pass

    docs = collection.find(
        query,
        {
            "_id": 0,
            "subject_code": 1,
            "subject_name": 1,
            "semester": 1,
            "year": 1,
        },
    )
    unique = {}

    for doc in docs:
        code = (doc.get("subject_code") or "").strip()
        name = (doc.get("subject_name") or "").strip()

        if not code or not name:
            continue

        key = (code, name, doc.get("semester"), doc.get("year"))
        unique[key] = {
            "subject_code": code,
            "subject_name": name,
            "semester": doc.get("semester"),
            "year": doc.get("year"),
            "label": f"{code} {name}",
        }

    return sorted(
        unique.values(),
        key=lambda item: (
            int(item.get("semester") or 0),
            str(item.get("subject_code") or ""),
            str(item.get("subject_name") or ""),
        ),
    )


def normalize_analytics_subject(subject):
    lowered = (subject or "").strip().lower()

    for canonical, aliases in ANALYTICS_SUBJECT_ALIASES.items():
        if lowered == canonical.lower():
            return canonical

        for alias in aliases:
            if lowered == alias or alias in lowered or lowered in alias:
                return canonical

    return (subject or "").strip()


def is_mumbai_university(university):
    lowered = (university or "").strip().lower()
    return "mumbai" in lowered


def normalize_analytics_university(university):
    lowered = (university or "").strip().lower()

    if not lowered:
        return ""
    if "sppu" in lowered or "savitribai" in lowered:
        return "SPPU"
    if "anna" in lowered:
        return "Anna University"
    if lowered == "iit" or "indian institute" in lowered:
        return "IIT"
    if "mumbai" in lowered:
        return "Mumbai University"

    return (university or "").strip()


def load_mumbai_course_outcomes_for_analytics(subject_code="", subject_name="", schema="C"):
    collection = get_subject_collection(schema)
    docs = []

    if subject_code:
        docs = list(
            collection.find(
                {"subject_code": {"$regex": f"^{re.escape(subject_code)}$", "$options": "i"}}
            ).sort("co_number", 1)
        )

    if not docs and subject_name:
        docs = list(
            collection.find(
                {"subject_name": {"$regex": f"^{re.escape(subject_name)}$", "$options": "i"}}
            ).sort("co_number", 1)
        )

    if not docs and subject_name:
        docs = list(
            collection.find(
                {"subject_name": {"$regex": re.escape(subject_name), "$options": "i"}}
            ).sort("co_number", 1)
        )

    if not docs:
        return None

    course_outcomes = [
        (doc.get("co_text") or "").strip()
        for doc in docs
        if (doc.get("co_text") or "").strip()
    ]

    if not course_outcomes:
        return None

    first_doc = docs[0]
    return {
        "subject": (first_doc.get("subject_name") or subject_name or subject_code or "").strip(),
        "subject_code": (first_doc.get("subject_code") or subject_code or "").strip(),
        "course_outcomes": course_outcomes,
        "source": normalize_schema_value(schema),
    }


def load_general_course_outcomes_for_analytics(university, subject):
    normalized_university = normalize_analytics_university(university)
    normalized_subject = normalize_analytics_subject(subject)

    data = db["university_co"].find_one({
        "university": normalized_university,
        "subject": normalized_subject
    })

    if not data:
        data = db["university_co"].find_one({
            "university": {"$regex": f"^{re.escape(normalized_university)}$", "$options": "i"},
            "subject": {"$regex": f"^{re.escape(normalized_subject)}$", "$options": "i"}
        })

    if not data:
        data = db["university_co"].find_one({
            "university": {"$regex": normalized_university, "$options": "i"},
            "subject": {"$regex": normalized_subject, "$options": "i"}
        })

    if not data:
        return None

    course_outcomes = [
        (co.get("description") or "").strip()
        for co in data.get("course_outcomes", [])
        if (co.get("description") or "").strip()
    ]

    if not course_outcomes:
        return None

    return {
        "subject": (data.get("subject") or normalized_subject).strip(),
        "subject_code": "",
        "course_outcomes": course_outcomes,
        "source": "university_co",
    }


def attach_wk_matches(course_outcomes, results, level_thresholds=None):
    wk_catalog = load_wk_catalog()

    if not wk_catalog or not results:
        return []

    wk_results = map_texts_to_candidates(
        texts=course_outcomes,
        candidates=wk_catalog,
        code_key="code",
        text_key="text",
        result_key="wk_code",
        level_thresholds=level_thresholds,
    )
    wk_lookup = {item["code"]: item["text"] for item in wk_catalog}

    for co_result, wk_result in zip(results, wk_results):
        wk_code = wk_result.get("wk_code")
        co_result["wk_match"] = {
            "wk_code": wk_code,
            "wk_text": wk_lookup.get(wk_code, ""),
            "confidence": wk_result.get("confidence", 0),
            "confidence_label": wk_result.get("confidence_label", "Low"),
            "level": wk_result.get("level", 1),
            "debug": wk_result.get("debug", {}),
        }

    return wk_catalog


def load_level_thresholds():
    now = time.monotonic()
    cached_value = _level_threshold_cache.get("value")

    if cached_value is not None and now < _level_threshold_cache.get("expires_at", 0.0):
        return dict(cached_value)

    default_thresholds = {
        "level_2_min": 0.40,
        "level_3_min": 0.60,
        "confidence_medium_min": 0.45,
        "confidence_high_min": 0.70,
        "sample_size": 0,
    }
    score_by_level = {1: [], 2: [], 3: []}

    for doc in mapping_collection.find(
        {"faculty_ratings": {"$exists": True, "$ne": []}},
        {"results": 1, "faculty_ratings": 1},
    ):
        score_lookup = {}

        for co_result in doc.get("results", []):
            co_code = co_result.get("co")

            for mapping in co_result.get("mapping", []):
                po_code = mapping.get("po_code")

                try:
                    score_lookup[(co_code, po_code)] = float(mapping.get("final_score", 0)) / 100.0
                except (TypeError, ValueError):
                    continue

        for rating in doc.get("faculty_ratings", []):
            if not isinstance(rating, dict):
                continue

            try:
                faculty_level = int(rating.get("faculty") or rating.get("faculty_level"))
            except (TypeError, ValueError):
                continue

            if faculty_level not in score_by_level:
                continue

            score = score_lookup.get((rating.get("co"), rating.get("po")))

            if score is not None:
                score_by_level[faculty_level].append(score)

    sample_size = sum(len(values) for values in score_by_level.values())

    if sample_size < 6:
        _level_threshold_cache["value"] = dict(default_thresholds)
        _level_threshold_cache["expires_at"] = now + RUNTIME_CACHE_TTL_SECONDS
        return dict(_level_threshold_cache["value"])

    means = {
        level: float(np.mean(values)) if values else None
        for level, values in score_by_level.items()
    }

    level_2_min = default_thresholds["level_2_min"]
    level_3_min = default_thresholds["level_3_min"]

    if means[1] is not None and means[2] is not None:
        level_2_min = (means[1] + means[2]) / 2.0
    elif means[2] is not None:
        level_2_min = means[2] - 0.10

    if means[2] is not None and means[3] is not None:
        level_3_min = (means[2] + means[3]) / 2.0
    elif means[3] is not None:
        level_3_min = means[3] - 0.08

    level_2_min = float(min(max(level_2_min, 0.28), 0.58))
    level_3_min = float(min(max(level_3_min, level_2_min + 0.12), 0.88))
    confidence_medium_min = float(min(max(level_2_min + 0.05, 0.35), 0.68))
    confidence_high_min = float(min(max(level_3_min + 0.05, confidence_medium_min + 0.10), 0.92))

    computed_thresholds = {
        "level_2_min": round(level_2_min, 4),
        "level_3_min": round(level_3_min, 4),
        "confidence_medium_min": round(confidence_medium_min, 4),
        "confidence_high_min": round(confidence_high_min, 4),
        "sample_size": sample_size,
    }
    _level_threshold_cache["value"] = dict(computed_thresholds)
    _level_threshold_cache["expires_at"] = now + RUNTIME_CACHE_TTL_SECONDS
    return dict(_level_threshold_cache["value"])


# =========================
# SIMPLE TEXT SIMILARITY
# =========================
def simple_similarity(text1, text2):

    words1 = set(text1.lower().split())
    words2 = set(text2.lower().split())

    common = words1.intersection(words2)

    return len(common) / max(len(words1), 1)


def tfidf_similarity(cos, pos):
    return [[simple_similarity(c, p) for p in pos] for c in cos]


def use_similarity(cos, pos):
    return [[simple_similarity(c, p) for p in pos] for c in cos]


def bert_similarity(cos, pos):
    return [[simple_similarity(c, p) for p in pos] for c in cos]


# =========================
# BLOOM LEVEL DETECTOR
# =========================
def get_bloom_level(text):

    text = text.lower()

    if re.search(r"define|list|identify", text):
        return "Remember"

    elif re.search(r"explain|describe", text):
        return "Understand"

    elif re.search(r"apply|implement|use", text):
        return "Apply"

    elif re.search(r"analyze|compare", text):
        return "Analyze"

    elif re.search(r"evaluate|justify", text):
        return "Evaluate"

    elif re.search(r"design|create|develop", text):
        return "Create"

    else:
        return "Understand"


# =========================
# AI REASON GENERATOR
# =========================
def generate_reason(co, po_text, score, bloom):

    co_clean = co.lower().replace("co", "").strip()

    templates = {

        "Remember":[
            f"Understanding basic concepts in '{co_clean}' supports the program outcome '{po_text}'.",
            f"This course outcome introduces fundamental knowledge required for '{po_text}'.",
        ],

        "Understand":[
            f"Explaining concepts related to '{co_clean}' contributes to achieving '{po_text}'.",
            f"This learning objective strengthens conceptual understanding aligned with '{po_text}'.",
        ],

        "Apply":[
            f"Applying the concept '{co_clean}' helps students develop competencies related to '{po_text}'.",
            f"Using techniques from '{co_clean}' supports the program outcome '{po_text}'.",
        ],

        "Analyze":[
            f"Analyzing problems in '{co_clean}' contributes to achieving the program outcome '{po_text}'.",
            f"This course outcome develops analytical skills associated with '{po_text}'.",
        ],

        "Evaluate":[
            f"Evaluating solutions related to '{co_clean}' supports '{po_text}'.",
            f"This outcome develops critical thinking abilities aligned with '{po_text}'.",
        ],

        "Create":[
            f"Designing solutions in '{co_clean}' directly contributes to the program outcome '{po_text}'.",
            f"This outcome enables students to create solutions aligned with '{po_text}'.",
        ]
    }

    if bloom in templates:
        return random.choice(templates[bloom])
    else:
        return f"The course outcome '{co_clean}' contributes to achieving the program outcome '{po_text}'."


# =========================
# ROOT API
# =========================
@app.get("/")
def home():
    return {"message": "AI NLP CO-PO Mapper Running 🚀"}


@app.get("/subject-catalog")
def subject_catalog(schema: str = "C", year: str = "", semester: str = ""):
    subjects = load_subject_catalog(schema=schema, year=year, semester=semester)
    return {
        "schema": normalize_schema_value(schema),
        "year": normalize_year_value(year),
        "semester": int(semester) if str(semester).strip().isdigit() else None,
        "subjects": subjects,
    }


@app.get("/university-subjects")
def university_subjects(university: str):
    cleaned_university = normalize_analytics_university(university)

    if not cleaned_university:
        raise HTTPException(status_code=400, detail="University is required")

    subjects = sorted(
        {
            value.strip()
            for value in db["university_co"].distinct(
                "subject",
                {"university": {"$regex": f"^{re.escape(cleaned_university)}$", "$options": "i"}},
            )
            if isinstance(value, str) and value.strip()
        }
    )

    return {
        "university": cleaned_university,
        "subjects": [{"value": subject, "label": subject} for subject in subjects],
    }


# =========================
# SIGNUP API
# =========================
@app.post("/signup")
def signup(data: dict):

    name = data.get("name")
    email = data.get("email")
    department = data.get("department")
    password = data.get("password")

    if users_collection.find_one({"email": email}):
        return {"status": "error", "msg": "User already exists"}

    if not password:
        return {"status": "error", "msg": "Password is required"}

    # Hash the password with bcrypt
    hashed_password = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

    users_collection.insert_one({
        "name": name,
        "email": email,
        "department": department,
        "password": hashed_password
    })

    return {"status": "success"}


@app.post("/detect-co")
def detect_co(data: dict):

    try:
        questions = data.get("questions")
        subject_code = data.get("subject")
        schema = data.get("schema", "C")   # 🔥 NEW LINE

        print("Subject Code:", subject_code)
        print("Schema:", schema)

        # 🔥 STEP 1: Fetch CO from DB (schema-based)
        cos_data = get_cos_by_subject(subject_code, schema)

        print("CO fetched:", len(cos_data))

        if not cos_data:
            return {
                "results": [
                    {
                        "question": q,
                        "co": "Not Found",
                        "confidence": 0,
                        "level": 0
                    } for q in questions
                ]
            }

        # 🔥 STEP 2: AI Mapping
        results = map_questions_to_co(questions, cos_data)

        return {"results": results}

    except Exception as e:
        print("ERROR:", e)
        raise HTTPException(status_code=500, detail=str(e))


# =========================
# CO ATTAINMENT API
# =========================

from fastapi import UploadFile, File, Form
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import shutil
import subprocess
import tempfile
from io import BytesIO


def _clean_extracted_co(text):
    cleaned = re.sub(r"\s+", " ", str(text or "")).strip(" :-–—\t\n")
    cleaned = re.sub(r"^(course\s*outcomes?|outcomes?)\s*[:\-]?\s*", "", cleaned, flags=re.I)
    cleaned = re.sub(r"^(course\s*objectives?|objectives?)\s*[:\-]?\s*", "", cleaned, flags=re.I)
    cleaned = re.sub(r"^(co|c0)\s*\d+\s*[:.)\-]?\s*", "", cleaned, flags=re.I)
    return cleaned.strip()


def _dedupe_text_items(items, limit=8):
    deduped = []
    seen = set()

    for item in items:
        cleaned = re.sub(r"\s+", " ", str(item or "")).strip(" :-–—\t\n")
        if len(cleaned) < 12:
            continue

        key = cleaned.lower()
        if key in seen:
            continue

        seen.add(key)
        deduped.append(cleaned)

        if limit and len(deduped) >= limit:
            break

    return deduped


def extract_course_outcomes_from_ocr_text(text):
    raw_text = str(text or "").replace("\r", "\n").replace("\u00a0", " ")
    raw_text = raw_text.replace("C O", "CO").replace("c o", "CO")
    raw_text = re.sub(r"\bC0\s*([1-9]\d?)\b", r"CO\1", raw_text, flags=re.I)
    raw_text = re.sub(r"\bCO\s*([Il])\b", "CO1", raw_text, flags=re.I)
    raw_text = re.sub(r"draft\s+copy", " ", raw_text, flags=re.I)
    raw_text = re.sub(r"[ \t]+", " ", raw_text)
    raw_text = re.sub(r"\n{2,}", "\n", raw_text)
    raw_text = re.sub(r"(?<!\n)\s+(?=(?:CO|C0)?\s*[0-9Il]{1,2}\s*[:.)\-])", "\n", raw_text, flags=re.I)
    raw_text = re.sub(r"(?<!\n)\s+(?=\d+\s*[:.)\-]\s+)", "\n", raw_text)

    stop_heading_pattern = re.compile(
        r"^\s*(module|content|unit|syllabus|prerequisite|teaching\s+scheme|credits?|examination\s+scheme|text\s*books?|references?|program\s*outcomes?|assessment|course\s*objectives?|course\s*code|course\s*title)\b",
        re.I,
    )
    numbered_item_pattern = re.compile(
        r"^\s*(?:CO|C0)?\s*([0-9Il]{1,2})\s*[:.)\-]?\s*(.*)$",
        re.I,
    )
    generic_numbered_item_pattern = re.compile(
        r"^\s*(\d+)\s*[:.)\-]?\s*(.*)$",
        re.I,
    )
    section_patterns = [
        ("Course Outcomes", re.compile(r"course\s*outcomes?\s*:?\s*(?:students\s+will\s+be\s+able\s+to)?", re.I)),
        ("Course Objectives", re.compile(r"course\s*objectives?\s*:?\s*(?:the\s+course\s+is\s+aimed\s+to)?", re.I)),
    ]
    stop_section_inline_pattern = re.compile(
        r"\b(module|content|unit|syllabus|prerequisite|teaching\s+scheme|credits?|examination\s+scheme|text\s*books?|references?|program\s*outcomes?|assessment|course\s*objectives?|course\s*code|course\s*title)\b",
        re.I,
    )

    def flush_item(buffer, target):
        if not buffer:
            return

        cleaned = _clean_extracted_co(" ".join(buffer))
        cleaned = re.sub(r"\s+", " ", cleaned).strip()

        if len(cleaned) >= 12:
            target.append(cleaned)

    lines = [re.sub(r"\s+", " ", line).strip(" |:-\t") for line in raw_text.split("\n")]

    def split_numbered_section(section_text):
        section_text = str(section_text or "").strip()
        if not section_text:
            return []

        section_text = re.sub(r"\s+", " ", section_text).strip(" :-")
        section_text = stop_section_inline_pattern.split(section_text, maxsplit=1)[0].strip()

        numbered_chunks = []
        numbered_pattern = re.compile(
            r"(?is)(?:^|\s)(?:CO|C0)?\s*([0-9Il]{1,2})\s*[:.)\-]?\s+(.*?)(?=(?:\s(?:CO|C0)?\s*[0-9Il]{1,2}\s*[:.)\-]?\s+)|\Z)"
        )

        for match in numbered_pattern.finditer(section_text):
            chunk = _clean_extracted_co(match.group(2))
            chunk = re.sub(r"\s+", " ", chunk).strip()
            if len(chunk) >= 12:
                numbered_chunks.append(chunk)

        if numbered_chunks:
            return _dedupe_text_items(numbered_chunks, limit=8)

        return []

    def extract_section_items(section_pattern):
        extracted_items = []
        collecting = False
        current_item = []

        for raw_line in lines:
            line = str(raw_line or "").strip()
            if not line:
                continue

            if not collecting:
                if section_pattern.search(line):
                    collecting = True
                    line = section_pattern.sub("", line, count=1).strip(" :-")
                    if not line or re.fullmatch(
                        r"(students will be able to|the student will be able to|the course is aimed to)\.?",
                        line,
                        flags=re.I,
                    ):
                        continue
                else:
                    continue

            if stop_heading_pattern.match(line):
                flush_item(current_item, extracted_items)
                current_item = []
                break

            if re.fullmatch(
                r"(students will be able to|the student will be able to|the course is aimed to)\.?",
                line,
                flags=re.I,
            ):
                continue

            numbered_match = numbered_item_pattern.match(line) or generic_numbered_item_pattern.match(line)
            if numbered_match:
                flush_item(current_item, extracted_items)
                current_item = []
                immediate_text = (numbered_match.group(2) or "").strip()
                if immediate_text:
                    current_item.append(immediate_text)
                continue

            if re.fullmatch(r"(?:CO|C0)?\s*[0-9Il]{1,2}", line, flags=re.I):
                flush_item(current_item, extracted_items)
                current_item = []
                continue

            current_item.append(line)

        flush_item(current_item, extracted_items)
        return _dedupe_text_items(extracted_items, limit=8)

    course_outcomes = []
    extracted_section = ""

    for section_name, section_pattern in section_patterns:
        direct_section_match = re.search(
            rf"{section_pattern.pattern}\s*(.*?)(?={stop_section_inline_pattern.pattern}|\Z)",
            raw_text,
            flags=re.I | re.S,
        )
        if direct_section_match:
            inline_items = split_numbered_section(direct_section_match.group(1))
            if inline_items:
                course_outcomes = inline_items
                extracted_section = section_name
                if section_name == "Course Outcomes":
                    break
                continue

        items = extract_section_items(section_pattern)
        if items:
            course_outcomes = items
            extracted_section = section_name
            if section_name == "Course Outcomes":
                break

    if not course_outcomes:
        co_pattern = re.compile(
            r"(?is)\bCO\s*[-:]?\s*([0-9Il]{1,2})\s*[:.)\-]?\s*(.*?)(?=\bCO\s*[-:]?\s*[0-9Il]{1,2}\s*[:.)\-]?|\b(program\s*outcomes?|text\s*books?|references?|module|syllabus|assessment|content)\b|\Z)"
        )

        for match in co_pattern.finditer(raw_text):
            co_text = _clean_extracted_co(match.group(2))
            if len(co_text) >= 12:
                course_outcomes.append(co_text)

    return {
        "items": _dedupe_text_items(course_outcomes, limit=8),
        "section_label": extracted_section if course_outcomes else "",
    }


def extract_text_from_pdf_bytes(file_bytes):
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError("pypdf is not installed.") from exc

    reader = PdfReader(BytesIO(file_bytes))
    chunks = []

    for page in reader.pages[:15]:
        page_text = page.extract_text() or ""
        if page_text.strip():
            chunks.append(page_text)

    return "\n".join(chunks).strip()


def _clean_syllabus_topic_line(line):
    cleaned = str(line or "").strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    cleaned = re.sub(r"^\s*(module|unit)\s*[-:]?\s*\d+\.?\d*\s*", "", cleaned, flags=re.I)
    cleaned = re.sub(r"^\s*\d+(\.\d+)?\s*", "", cleaned)
    cleaned = re.sub(r"\b\d+\s*$", "", cleaned)
    cleaned = cleaned.strip(" :-–—|")

    noise_patterns = [
        r"^hours?$",
        r"^marks?$",
        r"^module$",
        r"^sr\.?\s*no\.?$",
        r"^contents?$",
        r"^references?$",
        r"^text\s*books?$",
        r"^assessment$",
        r"^page\s+\d+",
    ]

    if any(re.search(pattern, cleaned, flags=re.I) for pattern in noise_patterns):
        return ""

    if len(cleaned) < 18:
        return ""

    alpha_ratio = sum(ch.isalpha() for ch in cleaned) / max(len(cleaned), 1)
    if alpha_ratio < 0.45:
        return ""

    return cleaned


def extract_syllabus_topics_from_ocr_text(text):
    raw_text = str(text or "").replace("\r", "\n")
    lines = [_clean_syllabus_topic_line(line) for line in raw_text.split("\n")]
    topics = [line for line in lines if line]

    if not topics:
        sentence_chunks = re.split(r"(?<=[.;])\s+|\n+", raw_text)
        topics = [_clean_syllabus_topic_line(chunk) for chunk in sentence_chunks]
        topics = [topic for topic in topics if topic]

    merged_topics = []

    for topic in topics:
        if merged_topics and len(topic) < 55 and not re.search(r"^[A-Z][A-Za-z\s,&/-]{8,}$", topic):
            merged_topics[-1] = f"{merged_topics[-1]} {topic}"
        else:
            merged_topics.append(topic)

    return _dedupe_text_items(merged_topics, limit=6)


def build_course_outcomes_from_syllabus_topics(topics):
    outcomes = []

    for topic in topics:
        lowered = topic.lower()

        if any(word in lowered for word in ["design", "develop", "implement", "create", "build"]):
            verb = "Design and implement"
        elif any(word in lowered for word in ["analyze", "analysis", "optimization", "regularization", "training"]):
            verb = "Analyze and apply"
        elif any(word in lowered for word in ["apply", "use", "model", "algorithm"]):
            verb = "Apply"
        else:
            verb = "Understand"

        topic_text = topic[0].lower() + topic[1:] if topic else topic
        outcomes.append(f"{verb} {topic_text}.")

    return _dedupe_text_items(outcomes, limit=6)


def ocr_image_with_easyocr(image):
    try:
        import easyocr
        import numpy as np
    except ImportError as exc:
        raise RuntimeError("EasyOCR is not installed.") from exc

    reader = easyocr.Reader(["en"], gpu=False, verbose=False)
    image_array = np.array(image)
    result = reader.readtext(image_array, detail=0, paragraph=True)
    return "\n".join(str(item) for item in result if str(item).strip())


@app.post("/extract-course-outcomes-image")
@app.post("/extract-course-outcomes-file")
async def extract_course_outcomes_image(file: UploadFile = File(...)):
    content_type = (file.content_type or "").lower()
    filename = (file.filename or "").lower()
    is_image = content_type.startswith("image/") or filename.endswith((".png", ".jpg", ".jpeg", ".webp", ".bmp"))
    is_pdf = "pdf" in content_type or filename.endswith(".pdf")

    if not is_image and not is_pdf:
        raise HTTPException(status_code=400, detail="Please upload a syllabus image or PDF file.")

    try:
        content = await file.read()
        raw_text = ""
        engine = ""
        source_type = "pdf" if is_pdf else "image"

        if is_pdf:
            engine = "pypdf"
            raw_text = extract_text_from_pdf_bytes(content)
        else:
            try:
                from PIL import Image
            except ImportError as exc:
                raise HTTPException(
                    status_code=500,
                    detail="Pillow is missing. Install backend requirements, then restart the backend.",
                ) from exc

            image = Image.open(BytesIO(content))
            image = image.convert("RGB")
            engine = "easyocr"
            raw_text = ocr_image_with_easyocr(image)

        extraction = extract_course_outcomes_from_ocr_text(raw_text)
        course_outcomes = extraction.get("items", [])

        return {
            "status": "success",
            "course_outcomes": course_outcomes,
            "raw_text": raw_text,
            "count": len(course_outcomes),
            "engine": engine,
            "source_type": source_type,
            "section_label": extraction.get("section_label", ""),
        }
    except HTTPException:
        raise
    except RuntimeError as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Unable to read the syllabus file. Install backend requirements and restart the backend. Error: {exc}",
        ) from exc
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Unable to extract course outcomes from the uploaded syllabus file: {exc}",
        ) from exc

def parse_cos(co_str):
    if pd.isna(co_str) or str(co_str).strip() == "":
        return []

    parts = re.findall(r"CO\d+", str(co_str), re.IGNORECASE)
    return [part.upper().strip() for part in parts]


def q_level(pct, scheme="C"):
    """
    C Scheme   : pct>=70->3, pct>=60->2, pct>=50->1, else 0
    NEP Scheme : pct>=60->3, pct>=50->2, pct>=40->1, else 0
    """
    normalized_scheme = (scheme or "C").upper()
    if normalized_scheme == "NEP":
        if pct >= 60:
            return 3
        if pct >= 50:
            return 2
        if pct >= 40:
            return 1
        return 0

    if pct >= 70:
        return 3
    if pct >= 60:
        return 2
    if pct >= 50:
        return 1
    return 0


def _find_header_row(df):
    markers = (
        "roll no.", "roll no", "roll number",
        "sr. no.", "sr. no", "sr no.", "sr no", "sr.no.", "sr.no",
        "serial no", "serial number", "s.no.", "s.no", "s. no",
        "seat no.", "seat no", "seat number"
    )

    for index, row in df.iterrows():
        row_text = " | ".join("" if pd.isna(value) else str(value) for value in row.values).lower()
        if any(marker in row_text for marker in markers):
            return index

    return None


def parse_attainment_data(xlsx_path, scheme="C"):
    scheme = (scheme or "C").upper()
    if scheme not in {"C", "NEP"}:
        scheme = "C"

    df = pd.read_excel(xlsx_path, sheet_name=0, header=None)
    header_row = _find_header_row(df)

    if header_row is None:
        raise ValueError("Could not find 'Roll No.' or 'Sr. No.' header in the uploaded file.")

    co_row = header_row - 3
    marks_row = header_row - 1
    ia_row = header_row - 4
    data_start = header_row + 1

    if co_row < 0 or marks_row < 0:
        raise ValueError("Uploaded file format is not valid for attainment analysis.")

    data_end = len(df)
    for row_index in range(data_start, len(df)):
        roll_value = df.iloc[row_index, 1] if df.shape[1] > 1 else None
        if pd.isna(roll_value) or str(roll_value).strip() == "":
            data_end = row_index
            break

    header_vals = df.iloc[header_row].tolist()
    marks_vals = df.iloc[marks_row].tolist()
    co_vals = df.iloc[co_row].tolist()
    ia_vals = df.iloc[ia_row].tolist() if ia_row >= 0 else []

    ia_map = {}
    current_ia = "IA1"
    for column_index, value in enumerate(ia_vals):
        if not pd.isna(value) and str(value).strip():
            label = str(value).strip().upper()
            current_ia = "IA2" if "2" in label else "IA1"
        ia_map[column_index] = current_ia

    questions = []
    for column_index, header_value in enumerate(header_vals):
        if str(header_value).strip().startswith("Q."):
            cos = parse_cos(co_vals[column_index])
            max_marks = marks_vals[column_index]

            if not pd.isna(max_marks) and cos:
                questions.append(
                    {
                        "col": column_index,
                        "label": str(header_value).strip(),
                        "cos": cos,
                        "max": float(max_marks),
                        "ia": ia_map.get(column_index, "IA1"),
                    }
                )

    if not questions:
        raise ValueError("No question columns with CO mappings were found in the uploaded file.")

    students = df.iloc[data_start:data_end].copy()
    total_students = len(students)
    all_cos = sorted({co for question in questions for co in question["cos"]})

    q_attainments = []
    for question in questions:
        threshold = question["max"] * 0.6
        column_data = pd.to_numeric(students.iloc[:, question["col"]], errors="coerce").dropna()
        appeared = len(column_data)
        above_threshold = int((column_data >= threshold).sum())
        pct = round((above_threshold / appeared * 100) if appeared else 0, 4)
        level = q_level(pct, scheme)

        q_attainments.append(
            {
                "label": question["label"],
                "cos": question["cos"],
                "max": question["max"],
                "threshold": round(threshold, 2),
                "appeared": appeared,
                "above_thr": above_threshold,
                "pct": round(pct, 2),
                "level": level,
                "ia": question.get("ia", "IA1"),
            }
        )

    co_attainments = {}
    for co in all_cos:
        levels = [question["level"] for question in q_attainments if co in question["cos"]]
        co_attainments[co] = round(sum(levels) / len(levels), 2) if levels else 0

    return {
        "total_students": total_students,
        "questions": questions,
        "q_attainments": q_attainments,
        "co_attainments": co_attainments,
        "all_cos": all_cos,
        "scheme": scheme,
    }


def build_attainment_excel(xlsx_path, result):
    workbook = load_workbook(xlsx_path)

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    green_fill = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    blue_fill = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")
    yellow_fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    orange_fill = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
    red_fill = PatternFill("solid", start_color="F4CCCC", end_color="F4CCCC")
    ia1_fill = PatternFill("solid", start_color="D6E4F7", end_color="D6E4F7")
    ia2_fill = PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if "CO Attainment Summary" in workbook.sheetnames:
        del workbook["CO Attainment Summary"]

    sheet = workbook.create_sheet("CO Attainment Summary", 1)

    for column, width in zip(["A", "B", "C", "D", "E", "F", "G"], [14, 22, 13, 16, 24, 14, 14]):
        sheet.column_dimensions[column].width = width

    def write_cell(row_number, column_number, value, fill=None, bold=False, white=False, alignment=None, number_format=None):
        cell = sheet.cell(row=row_number, column=column_number, value=value)
        cell.font = Font(
            bold=bold,
            name="Arial",
            size=10,
            color="FFFFFF" if white else "000000",
        )
        cell.alignment = alignment or center
        cell.border = border
        if number_format:
            cell.number_format = number_format
        if fill:
            cell.fill = fill

    original_sheet = workbook[workbook.sheetnames[0]]

    def write_original_cell(row_number, column_number, value, fill=None, bold=False, white=False, alignment=None, number_format=None):
        cell = original_sheet.cell(row=row_number, column=column_number, value=value)
        cell.font = Font(
            bold=bold,
            name="Arial",
            size=11,
            color="FFFFFF" if white else "000000",
        )
        cell.alignment = alignment or center
        cell.border = border
        if number_format:
            cell.number_format = number_format
        if fill:
            cell.fill = fill

    row_number = 1
    sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=7)
    write_cell(
        row_number,
        1,
        "CO Attainment Report - Per Question Analysis",
        header_fill,
        bold=True,
        white=True,
    )
    sheet.row_dimensions[row_number].height = 24
    row_number += 1

    headers = [
        "Question",
        "Mapped COs",
        "Max Marks",
        "60% Threshold",
        "Students >= Threshold",
        "% Above 60%",
        "Attainment Level",
    ]
    for column_number, header in enumerate(headers, 1):
        write_cell(row_number, column_number, header, blue_fill, bold=True)
    row_number += 1

    for question in result["q_attainments"]:
        level = question["level"]
        fill = (
            green_fill if level == 3 else
            yellow_fill if level == 2 else
            orange_fill if level == 1 else
            red_fill
        )

        write_cell(row_number, 1, question["label"])
        write_cell(row_number, 2, ", ".join(question["cos"]), alignment=left)
        write_cell(row_number, 3, question["max"])
        write_cell(row_number, 4, question["threshold"])
        write_cell(row_number, 5, f"{question['above_thr']} / {question['appeared']}")
        write_cell(row_number, 6, f"{question['pct']}%")
        write_cell(row_number, 7, f"Level {level}", fill)
        row_number += 1

    row_number += 1
    horizontal_start_row = row_number
    horizontal_last_col = len(result["q_attainments"]) + 1

    for column_index in range(2, horizontal_last_col + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 14
    sheet.column_dimensions["A"].width = max(sheet.column_dimensions["A"].width, 38)

    sheet.merge_cells(
        start_row=horizontal_start_row,
        start_column=1,
        end_row=horizontal_start_row,
        end_column=horizontal_last_col,
    )
    write_cell(
        horizontal_start_row,
        1,
        "Question-wise CO Attainment Summary",
        header_fill,
        bold=True,
        white=True,
    )
    sheet.row_dimensions[horizontal_start_row].height = 24

    def write_question_metric(row_number, label, values, fill=None, label_fill=None):
        write_cell(row_number, 1, label, fill=label_fill, bold=True, alignment=left)
        for offset, value in enumerate(values, start=2):
            write_cell(row_number, offset, value, fill=fill)

    def ia_fill_for(question):
        return ia1_fill if question.get("ia", "IA1") == "IA1" else ia2_fill

    header_row = horizontal_start_row + 1
    write_question_metric(
        header_row,
        "Internal Assessment",
        [question.get("ia", "IA1") for question in result["q_attainments"]],
        label_fill=blue_fill,
    )
    for offset, question in enumerate(result["q_attainments"], start=2):
        write_cell(header_row, offset, question.get("ia", "IA1"), fill=ia_fill_for(question), bold=True)

    write_question_metric(
        header_row + 1,
        "Question",
        [question["label"] for question in result["q_attainments"]],
        fill=blue_fill,
        label_fill=blue_fill,
    )
    write_question_metric(
        header_row + 2,
        "Number of students appeared",
        [question["appeared"] for question in result["q_attainments"]],
    )
    write_question_metric(
        header_row + 3,
        "60% of respective question",
        [question["threshold"] for question in result["q_attainments"]],
    )
    write_question_metric(
        header_row + 4,
        "No. of students having more than 60% marks in respective question",
        [question["above_thr"] for question in result["q_attainments"]],
    )
    write_question_metric(
        header_row + 5,
        f"Percentage of students above 60% (Total={result['total_students']})",
        [round(question["pct"]) for question in result["q_attainments"]],
    )
    write_cell(header_row + 6, 1, "CO Attainment out of 3", bold=True, alignment=left)
    for offset, question in enumerate(result["q_attainments"], start=2):
        level = question["level"]
        fill = (
            green_fill if level == 3 else
            yellow_fill if level == 2 else
            orange_fill if level == 1 else
            red_fill
        )
        write_cell(header_row + 6, offset, level, fill=fill, bold=True)

    write_cell(header_row + 7, 1, "Mapped COs", bold=True, alignment=left)
    sheet.row_dimensions[header_row + 7].height = 42
    for offset, question in enumerate(result["q_attainments"], start=2):
        write_cell(
            header_row + 7,
            offset,
            ", ".join(question["cos"]),
            fill=yellow_fill,
            bold=True,
        )

    average_start_row = header_row + 9
    average_last_col = len(result["all_cos"]) + 1

    sheet.merge_cells(
        start_row=average_start_row,
        start_column=1,
        end_row=average_start_row,
        end_column=average_last_col,
    )
    write_cell(
        average_start_row,
        1,
        "CO Attainment (Average)",
        header_fill,
        bold=True,
        white=True,
    )
    sheet.row_dimensions[average_start_row].height = 22

    write_cell(average_start_row + 1, 1, "CO", fill=blue_fill, bold=True)
    for offset, co in enumerate(result["all_cos"], start=2):
        write_cell(average_start_row + 1, offset, co, fill=blue_fill, bold=True)

    write_cell(average_start_row + 2, 1, "CO Attainment through PT", fill=yellow_fill, bold=True, alignment=left)
    for offset, co in enumerate(result["all_cos"], start=2):
        attainment_value = result["co_attainments"][co]
        fill = (
            green_fill if attainment_value >= 2.5 else
            yellow_fill if attainment_value >= 1.5 else
            orange_fill
        )
        write_cell(
            average_start_row + 2,
            offset,
            attainment_value,
            fill=fill,
            bold=True,
            number_format="0.00",
        )

    row_number = average_start_row + 4
    sheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=5)
    write_cell(
        row_number,
        1,
        "CO Attainment Summary (AVERAGE of question levels per CO)",
        header_fill,
        bold=True,
        white=True,
    )
    sheet.row_dimensions[row_number].height = 22
    row_number += 1

    summary_headers = ["CO", "Attainment Value", "Level (out of 3)", "Status", "Questions Mapped"]
    for column_number, header in enumerate(summary_headers, 1):
        write_cell(row_number, column_number, header, blue_fill, bold=True)
    row_number += 1

    for co in result["all_cos"]:
        attainment_value = result["co_attainments"][co]
        level = int(round(attainment_value))
        status = "High" if attainment_value >= 2.5 else "Medium" if attainment_value >= 1.5 else "Low"
        fill = green_fill if status == "High" else yellow_fill if status == "Medium" else orange_fill
        mapped_questions = [question["label"] for question in result["q_attainments"] if co in question["cos"]]

        write_cell(row_number, 1, co, fill=fill, bold=True)
        write_cell(row_number, 2, attainment_value, fill=fill)
        write_cell(row_number, 3, level, fill=fill)
        write_cell(row_number, 4, status, fill=fill)
        write_cell(row_number, 5, ", ".join(mapped_questions), fill=fill, alignment=left)
        row_number += 1

    row_number += 2
    write_cell(row_number, 1, "Total Students Analysed", bold=True)
    write_cell(row_number, 2, result["total_students"])
    row_number += 1
    write_cell(row_number, 1, "Formula Used", bold=True)
    sheet.merge_cells(start_row=row_number, start_column=2, end_row=row_number, end_column=5)
    formula_text = (
        "Level: pct>=60->3, pct>=50->2, pct>=40->1, else 0 | CO = AVERAGE(question levels)"
        if result.get("scheme") == "NEP"
        else "Level: pct>=70->3, pct>=60->2, pct>=50->1, else 0 | CO = AVERAGE(question levels)"
    )
    write_cell(
        row_number,
        2,
        formula_text,
    )

    def find_last_content_row(worksheet):
        for row_number in range(worksheet.max_row, 0, -1):
            if any(cell.value not in (None, "") for cell in worksheet[row_number]):
                return row_number
        return 1

    question_columns = [question["col"] + 1 for question in result["questions"]]
    if question_columns:
        start_row = find_last_content_row(original_sheet) + 2
        label_start_col = 2
        first_question_col = min(question_columns)
        label_end_col = max(label_start_col, first_question_col - 1)

        def merge_label(row_number, text, fill=None, bold=False):
            if label_end_col > label_start_col:
                original_sheet.merge_cells(
                    start_row=row_number,
                    start_column=label_start_col,
                    end_row=row_number,
                    end_column=label_end_col,
                )
            write_original_cell(
                row_number,
                label_start_col,
                text,
                fill=fill,
                bold=bold,
                alignment=left,
            )

        merge_label(start_row, "Number of students appeared", bold=True)
        merge_label(start_row + 1, "60% of respective question", bold=True)
        merge_label(start_row + 2, "No. of students having more than 60% marks in respective question", bold=True)
        merge_label(start_row + 3, f"Percentage of students above 60% (Total={result['total_students']})", bold=True)
        merge_label(start_row + 4, "CO Attainment out of 3", bold=True)

        for question, excel_col in zip(result["q_attainments"], question_columns):
            level = question["level"]
            level_fill = (
                green_fill if level == 3 else
                yellow_fill if level == 2 else
                orange_fill if level == 1 else
                red_fill
            )
            write_original_cell(start_row, excel_col, question["appeared"])
            write_original_cell(start_row + 1, excel_col, question["threshold"])
            write_original_cell(start_row + 2, excel_col, question["above_thr"])
            write_original_cell(start_row + 3, excel_col, round(question["pct"]))
            write_original_cell(start_row + 4, excel_col, level, fill=level_fill, bold=True)

        mapped_row = start_row + 5
        for question, excel_col in zip(result["q_attainments"], question_columns):
            write_original_cell(
                mapped_row,
                excel_col,
                ", ".join(question["cos"]),
                fill=yellow_fill,
                bold=True,
            )
        original_sheet.row_dimensions[mapped_row].height = 42

        average_header_row = mapped_row + 2
        merge_label(average_header_row, "CO Attainment(Average)", bold=True)
        for offset, co in enumerate(result["all_cos"], start=5):
            write_original_cell(average_header_row, offset, co, bold=True)

        average_value_row = average_header_row + 1
        merge_label(average_value_row, "CO Attainment through PT", fill=yellow_fill, bold=True)
        for offset, co in enumerate(result["all_cos"], start=5):
            attainment_value = result["co_attainments"][co]
            fill = (
                green_fill if attainment_value >= 2.5 else
                yellow_fill if attainment_value >= 1.5 else
                orange_fill
            )
            write_original_cell(
                average_value_row,
                offset,
                attainment_value,
                fill=fill,
                bold=True,
                number_format="0.00",
            )

    output_path = xlsx_path.replace(".xlsx", "_CO_Attainment_Report.xlsx")
    workbook.save(output_path)
    return output_path


def detect_excel_format(file_bytes, filename):
    lowered_name = (filename or "").lower()
    header = file_bytes[:8]

    if header.startswith(b"PK\x03\x04"):
        return ".xlsx"
    if header.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return ".xls"
    if lowered_name.endswith(".xls"):
        return ".xls"

    return ".xlsx"


def get_xlsx_path(file_bytes, filename, prefix):
    suffix = detect_excel_format(file_bytes, filename)
    temp_file = tempfile.NamedTemporaryFile(suffix=suffix, delete=False, prefix=prefix)
    temp_file.write(file_bytes)
    temp_file.close()

    if suffix == ".xlsx":
        return temp_file.name, temp_file.name

    converted_path = temp_file.name.replace(".xls", ".xlsx")
    office_binary = shutil.which("libreoffice") or shutil.which("soffice")

    if office_binary:
        process = subprocess.run(
            [office_binary, "--headless", "--convert-to", "xlsx", "--outdir", os.path.dirname(temp_file.name), temp_file.name],
            capture_output=True,
            text=True,
        )
        if process.returncode == 0 and os.path.exists(converted_path):
            return converted_path, temp_file.name

    try:
        excel_file = pd.ExcelFile(temp_file.name)
        with pd.ExcelWriter(converted_path, engine="openpyxl") as writer:
            for sheet_name in excel_file.sheet_names:
                pd.read_excel(temp_file.name, sheet_name=sheet_name, header=None).to_excel(
                    writer,
                    sheet_name=sheet_name[:31],
                    index=False,
                    header=False,
                )
        return converted_path, temp_file.name
    except Exception as exc:
        raise ValueError(
            "Unable to process .xls file. Please upload the file in .xlsx format."
        ) from exc


def _cleanup_temp_file(path):
    if path and os.path.exists(path):
        try:
            os.unlink(path)
        except OSError:
            pass


@app.post("/calculate-attainment")
async def calculate_attainment(file: UploadFile = File(...), scheme: str = Form("C")):
    source_path = None
    xlsx_path = None

    try:
        content = await file.read()
        xlsx_path, source_path = get_xlsx_path(content, file.filename or "attainment.xlsx", "attainment_")
        result = parse_attainment_data(xlsx_path, scheme=scheme.upper())
        report_path = build_attainment_excel(xlsx_path, result)

        if source_path != report_path:
            _cleanup_temp_file(source_path)
        if xlsx_path not in {report_path, source_path}:
            _cleanup_temp_file(xlsx_path)

        return FileResponse(
            path=report_path,
            filename="CO_Attainment_Report.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except HTTPException:
        raise
    except Exception as e:
        print("ATTAINMENT ERROR:", e)
        _cleanup_temp_file(source_path)
        if xlsx_path and xlsx_path != source_path:
            _cleanup_temp_file(xlsx_path)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/calculate-attainment-preview")
async def calculate_attainment_preview(file: UploadFile = File(...), scheme: str = Form("C")):
    source_path = None
    xlsx_path = None

    try:
        content = await file.read()
        xlsx_path, source_path = get_xlsx_path(content, file.filename or "preview.xlsx", "preview_")
        result = parse_attainment_data(xlsx_path, scheme=scheme.upper())

        return {
            "total_students": result["total_students"],
            "co_attainments": result["co_attainments"],
            "q_attainments": result["q_attainments"],
            "scheme": result["scheme"],
        }
    except HTTPException:
        raise
    except Exception as e:
        print("PREVIEW ERROR:", e)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        _cleanup_temp_file(source_path)
        if xlsx_path and xlsx_path != source_path:
            _cleanup_temp_file(xlsx_path)


# =========================
# LOGIN API
# =========================
@app.post("/login")
def login(data: dict):

    email = data.get("email")
    password = data.get("password")

    user = users_collection.find_one({"email": email})

    if not user:
        return {"status": "error", "msg": "User not found"}

    stored_password = user.get("password")
    if not password or not stored_password:
        return {"status": "error", "msg": "Invalid credentials"}

    try:
        is_correct = bcrypt.checkpw(password.encode("utf-8"), stored_password.encode("utf-8"))
    except Exception:
        # Fallback to plain-text verification for legacy accounts
        is_correct = (password == stored_password)
        if is_correct:
            # Upgrade stored password to a secure bcrypt hash
            try:
                hashed_password = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
                users_collection.update_one({"_id": user["_id"]}, {"$set": {"password": hashed_password}})
            except Exception as e:
                print("Failed to upgrade password hash for user:", email, e)

    if not is_correct:
        return {"status": "error", "msg": "Invalid credentials"}

    return {
        "status": "success",
        "name": user["name"]
    }


# =========================
# CO-PO MAPPING API
# =========================
from bson import ObjectId

# =========================
# MAP API (FINAL)
# =========================
@app.post("/map")
def map_copo(data: dict):

    try:
        # =========================
        # INPUT
        # =========================
        cos = [
            co.strip()
            for co in data.get("cos", [])
            if isinstance(co, str) and co.strip()
        ]
        mode = data.get("mode", "AICTE")

        if not cos:
            raise HTTPException(status_code=400, detail="At least one CO is required")

        # =========================
        # PO SELECT
        # =========================
        if mode == "AICTE":
            po_catalog = AICTE_POS
        else:
            custom_pos = [
                po.strip()
                for po in data.get("pos", [])
                if isinstance(po, str) and po.strip()
            ]

            if not custom_pos:
                raise HTTPException(
                    status_code=400,
                    detail="At least one custom PO is required for custom mode",
                )

            po_catalog = [
                {"code": f"PO{index + 1}", "text": text}
                for index, text in enumerate(custom_pos)
            ]

        faculty_name = (data.get("faculty_name") or "").strip()
        course_name = (data.get("course_name") or "").strip()
        course_code = (data.get("course_code") or "").strip()
        semester = (data.get("semester") or "").strip()
        subject = (data.get("subject") or "").strip()
        subject_name = subject or course_name

        level_thresholds = load_level_thresholds()
        results = map_course_outcomes_to_pos(
            course_outcomes=cos,
            po_catalog=po_catalog,
            top_k=len(po_catalog),
            level_thresholds=level_thresholds,
            subject_name=subject_name,
        )
        wk_catalog = attach_wk_matches(cos, results, level_thresholds=level_thresholds)
        course_outcomes_input_raw = data.get("course_outcomes_input_raw") or ""
        custom_pos_input_raw = data.get("custom_pos_input_raw") or ""
        generated_at = data.get("generated_at") or ""

        doc = {
            "faculty_name": faculty_name,
            "course_name": course_name,
            "course_code": course_code,
            "semester": semester,
            "subject": subject,
            "mode": mode,
            "course_outcomes": cos,
            "custom_pos": [
                po.strip()
                for po in data.get("pos", [])
                if isinstance(po, str) and po.strip()
            ],
            "course_outcomes_input_raw": course_outcomes_input_raw,
            "custom_pos_input_raw": custom_pos_input_raw,
            "generated_at": generated_at,
            "po_catalog": po_catalog,
            "wk_catalog": wk_catalog,
            "calibration": level_thresholds,
            "results": results,
        }

        inserted = mapping_collection.insert_one(doc)

        return {
            "id": str(inserted.inserted_id),
            "faculty_name": faculty_name,
            "course_name": course_name,
            "course_code": course_code,
            "semester": semester,
            "subject": subject,
            "mode": mode,
            "course_outcomes": cos,
            "custom_pos": doc["custom_pos"],
            "course_outcomes_input_raw": course_outcomes_input_raw,
            "custom_pos_input_raw": custom_pos_input_raw,
            "generated_at": generated_at,
            "po_catalog": po_catalog,
            "wk_catalog": wk_catalog,
            "calibration": level_thresholds,
            "results": results,
        }

        results = []

        # =========================
        # MAIN LOOP
        # =========================
        for i, co in enumerate(cos):

            bloom = get_bloom_level(co)

            # 🔥 STEP 1: SIMILARITY CALCULATE
            scores = []

            for j in range(len(pos_texts)):
                sim = simple_similarity(co, pos_texts[j])
                scores.append((j, sim))

            # 🔥 STEP 2: SORT (BEST MATCH FIRST)
            scores.sort(key=lambda x: x[1], reverse=True)

            # 🔥 STEP 3: TOP 3 SELECT
            top = scores[:3]

            mapping = []

            # 🔥 STEP 4: BUILD RESULT
            for j, sim in top:

                confidence = round(85+ (sim * 20), 2)
                confidence = min(confidence, 95)

                if confidence < 85:
                    level = 1
                elif confidence < 87:
                    level = 2
                else:
                    level = 3

                reason_text = generate_reason(co, pos_texts[j], confidence, bloom)

                mapping.append({
                    "po_code": pos_codes[j],
                    "po_text": pos_texts[j],
                    "final_score": confidence,
                    "level": level,
                    "reason": reason_text,
                    "justification": reason_text
                })

            results.append({
                "co": f"CO{i+1}",
                "bloom": bloom,
                "mapping": mapping
            })

        # =========================
        # 🔥 SAVE TO MONGODB
        # =========================
        doc = {
            "course_name": data.get("course_name"),
            "course_code": data.get("course_code"),
            "semester": data.get("semester"),
            "subject": data.get("subject"),
            "mode": mode,
            "results": results
        }

        inserted = mapping_collection.insert_one(doc)

        # =========================
        # RESPONSE
        # =========================
        return {
            "id": str(inserted.inserted_id),   # 🔥 IMPORTANT
            "results": results
        }

    except HTTPException:
        raise
    except Exception as e:
        print("MAP ERROR:", e)
        raise HTTPException(status_code=500, detail=str(e))

# =========================
# UPDATE JUSTIFICATION
# =========================
@app.post("/update-justification")
def update_justification(data: dict):
    justification_text = (
        data.get("justification")
        or data.get("reason")
        or data.get("justification_text")
        or data.get("reason_text")
        or ""
    )
    if not justification_text:
        raise HTTPException(status_code=400, detail="justification is required")

    selected_index = data.get("selected_justification_index")
    if selected_index is None:
        selected_index = data.get("selected_reason_index")

    justification_options = data.get("justification_options")
    if not isinstance(justification_options, list):
        justification_options = data.get("reason_options")

    update_fields = {
        "results.$[co].mapping.$[m].reason": justification_text,
        "results.$[co].mapping.$[m].justification": justification_text,
    }

    if selected_index is not None:
        update_fields["results.$[co].mapping.$[m].selected_reason_index"] = int(
            selected_index
        )
        update_fields["results.$[co].mapping.$[m].selected_justification_index"] = int(
            selected_index
        )

    if isinstance(justification_options, list):
        update_fields["results.$[co].mapping.$[m].reason_options"] = justification_options
        update_fields["results.$[co].mapping.$[m].justification_options"] = justification_options

    mapping_collection.update_one(
        {"_id": ObjectId(data["id"])},
        {
            "$set": update_fields
        },
        array_filters=[
            {"co.co": data["co"]},
            {"m.po_code": data["po"]}
        ]
    )

    return {"status": "updated"}

# =========================
# SAVE FACULTY RATINGS
# =========================
@app.post("/save-faculty")
def save_faculty(data: dict):

    try:
        id = data.get("id")
        ratings = data.get("ratings")

        if not id:
            return {"status": "error", "msg": "ID missing"}

        # 🔥 SAVE in SAME DOCUMENT
        mapping_collection.update_one(
            {"_id": ObjectId(id)},
            {
                "$set": {
                    "faculty_ratings": ratings
                }
            }
        )
        _level_threshold_cache["value"] = None
        _level_threshold_cache["expires_at"] = 0.0

        return {"status": "success"}

    except Exception as e:
        print("Faculty Save Error:", e)
        return {"status": "error"}


# =========================
# SAVE FEEDBACK
# =========================
@app.post("/feedback")
def save_feedback(data: dict):
    try:
        co = data.get("co")
        po = data.get("po")
        faculty_level = int(data.get("faculty_level"))
        ai_level = int(data.get("ai_level"))

        if not co or not po:
            raise HTTPException(status_code=400, detail="CO and PO are required")

        store_feedback(co, po, faculty_level, ai_level)

        db["feedback"].insert_one(
            {
                "co": co,
                "po": po,
                "faculty_level": faculty_level,
                "ai_level": ai_level,
            }
        )

        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        print("Feedback Save Error:", e)
        raise HTTPException(status_code=500, detail=str(e))


# =========================
# EVALUATE ACCURACY
# =========================
@app.post("/evaluate")
def evaluate_accuracy(data: dict):
    predicted = data.get("predicted", {})
    actual = data.get("actual", {})

    total = 0
    correct = 0

    for co, po_map in actual.items():
        for po, faculty_level in po_map.items():
            predicted_level = predicted.get(co, {}).get(po)

            if predicted_level is None:
                continue

            total += 1

            if int(predicted_level) == int(faculty_level):
                correct += 1

    accuracy = round((correct / total) * 100, 2) if total else 0.0

    return {
        "accuracy": accuracy,
        "correct": correct,
        "total": total,
    }


# =========================
# GET MAPPING
# =========================
@app.get("/get-mapping/{id}")
def get_mapping(id: str):

    data = mapping_collection.find_one({"_id": ObjectId(id)})

    data["_id"] = str(data["_id"])

    return data

# =========================
# ADD SUBJECT
# =========================
from fastapi import UploadFile, File, Form
from bson import ObjectId
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")

# folder create if not exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)


# 🔥 GET SUBJECTS
@app.get("/subjects")
def get_subjects():
    data = list(db["subjects"].find())

    for d in data:
        d["_id"] = str(d["_id"])

    return data


# 🔥 ADD SUBJECT (FILE SAVE)
@app.post("/add-subject")
async def add_subject(
    subjectName: str = Form(...),
    semester: str = Form(...),
    file: UploadFile = File(None)
):

    filename = ""

    if file:
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)

        with open(filepath, "wb") as f:
            f.write(await file.read())

        filename = file.filename

    data = {
        "subjectName": subjectName,
        "semester": semester,
        "file": filename
    }

    db["subjects"].insert_one(data)

    return {"message": "Saved successfully"}


# =========================
# FILE PREVIEW
# =========================
@app.get("/file/{filename}")
def get_file(filename: str):
    safe_name = os.path.basename(filename)
    path = os.path.join(UPLOAD_FOLDER, safe_name)

    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(path, filename=safe_name)


# 🔥 DELETE SUBJECT
from bson import ObjectId

@app.delete("/delete-subject/{id}")
def delete_subject(id: str):
    result = db["subjects"].delete_one({"_id": ObjectId(id)})

    if result.deleted_count == 1:
        return {"message": "Deleted successfully"}
    else:
        return {"message": "Not found"}


# =========================
# FETCH CO DATA
# =========================
# =========================
# ANALYZE API (FINAL)
# =========================
# =========================
# FINAL ANALYZE API (100% WORKING)
# =========================
@app.get("/analyze")
def analyze(university: str, subject: str = "", schema: str = "", subject_code: str = ""):
    try:
        cleaned_university = (university or "").strip()
        cleaned_subject = (subject or "").strip()
        cleaned_subject_code = (subject_code or "").strip()
        normalized_schema = normalize_schema_value(schema)

        if not cleaned_university:
            return {"status": "error", "msg": "University is required"}

        if is_mumbai_university(cleaned_university):
            source_data = load_mumbai_course_outcomes_for_analytics(
                subject_code=cleaned_subject_code,
                subject_name=cleaned_subject,
                schema=normalized_schema,
            )
        else:
            source_data = load_general_course_outcomes_for_analytics(
                university=cleaned_university,
                subject=cleaned_subject,
            )

        if not source_data:
            return {
                "status": "error",
                "msg": "CO data not found for the selected university and subject"
            }

        course_outcomes = source_data["course_outcomes"]

        if not course_outcomes:
            return {
                "status": "error",
                "msg": "No CO found for the selected subject"
            }

        po_docs = list(db["pos"].find({}, {"_id": 0, "code": 1, "description": 1}))
        po_catalog = [
            {"code": po.get("code"), "text": po.get("description")}
            for po in po_docs
            if po.get("code") and po.get("description")
        ]

        if not po_catalog:
            po_catalog = [
                {"code": po.get("code"), "text": po.get("text")}
                for po in AICTE_POS
                if po.get("code") and po.get("text")
            ]

        if not po_catalog:
            return {"status": "error", "msg": "PO data missing"}

        subject_name = source_data.get("subject", cleaned_subject) or cleaned_subject
        level_thresholds = load_level_thresholds()
        results = map_course_outcomes_to_pos(
            course_outcomes=course_outcomes,
            po_catalog=po_catalog,
            top_k=len(po_catalog),
            level_thresholds=level_thresholds,
            subject_name=subject_name,
        )
        wk_catalog = attach_wk_matches(course_outcomes, results, level_thresholds=level_thresholds)

        return {
            "status": "success",
            "university": cleaned_university,
            "subject": source_data.get("subject", cleaned_subject),
            "subject_code": source_data.get("subject_code", cleaned_subject_code),
            "source": source_data.get("source", "unknown"),
            "po_catalog": po_catalog,
            "wk_catalog": wk_catalog,
            "calibration": level_thresholds,
            "results": results
        }

        pos = [p["description"] for p in po_docs]
        po_codes = [p["code"] for p in po_docs]

        results = []

        # =========================
        # 🔥 CO → PO MAPPING
        # =========================
        for i, co in enumerate(cos):

            co_text = co.get("description", "")
            scores = []

            for j, po in enumerate(pos):

                words1 = set(co_text.lower().split())
                words2 = set(po.lower().split())

                common = words1.intersection(words2)
                score = len(common) / max(len(words1), 1)

                scores.append((j, score))

            scores.sort(key=lambda x: x[1], reverse=True)
            top = scores[:3]

            mapping = []

            for j, sc in top:
                level = 1 if sc < 0.3 else 2 if sc < 0.6 else 3

                mapping.append({
                    "po_code": po_codes[j],
                    "po_text": pos[j],
                    "level": level
                })

            results.append({
                "co": f"CO{i+1}",
                "objective": co_text,
                "bloom": "Understand",
                "mapping": mapping
            })

        return {
            "status": "success",
            "results": results
        }

    except Exception as e:
        print("ERROR:", e)
        return {
            "status": "error",
            "msg": str(e)
        }
